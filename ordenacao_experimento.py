import random
import time
import statistics
import multiprocessing as mp
import pandas as pd
import matplotlib.pyplot as plt
import os
import shutil

# ==========================
# CONFIGURAÇÕES DO TESTE
# ==========================

TAMANHOS = [1000, 10000, 100000]
ALGORITMOS = ["Insertion Sort", "Merge Sort", "Heap Sort"]
REPETICOES = 3
LIMITE_TEMPO = 300  # 5 minutos em segundos
SEED = 2026

ARQUIVO_MODELO = "PlanilhaListaDeExerciciosX.xlsx"
ARQUIVO_SAIDA = "PlanilhaListaDeExerciciosX_preenchida.xlsx"
ARQUIVO_GRAFICO = "grafico_tempos_medios.png"


# ==========================
# ALGORITMOS DE ORDENAÇÃO
# ==========================

def insertion_sort(vetor):
    movimentacoes = 0

    for i in range(1, len(vetor)):
        chave = vetor[i]
        movimentacoes += 1

        j = i - 1

        while j >= 0 and vetor[j] > chave:
            vetor[j + 1] = vetor[j]
            movimentacoes += 1
            j -= 1

        vetor[j + 1] = chave
        movimentacoes += 1

    return movimentacoes


def merge_sort(vetor):
    temp = [0] * len(vetor)

    def ordenar(inicio, fim):
        if fim - inicio <= 1:
            return 0

        meio = (inicio + fim) // 2

        movimentacoes = 0
        movimentacoes += ordenar(inicio, meio)
        movimentacoes += ordenar(meio, fim)

        i = inicio
        j = meio
        k = inicio

        while i < meio and j < fim:
            if vetor[i] <= vetor[j]:
                temp[k] = vetor[i]
                i += 1
            else:
                temp[k] = vetor[j]
                j += 1
            k += 1

        while i < meio:
            temp[k] = vetor[i]
            i += 1
            k += 1

        while j < fim:
            temp[k] = vetor[j]
            j += 1
            k += 1

        for pos in range(inicio, fim):
            vetor[pos] = temp[pos]
            movimentacoes += 1

        return movimentacoes

    return ordenar(0, len(vetor))


def heap_sort(vetor):
    trocas = 0

    def trocar(i, j):
        nonlocal trocas
        vetor[i], vetor[j] = vetor[j], vetor[i]
        trocas += 1

    def heapify(tamanho, raiz):
        maior = raiz
        esquerda = 2 * raiz + 1
        direita = 2 * raiz + 2

        if esquerda < tamanho and vetor[esquerda] > vetor[maior]:
            maior = esquerda

        if direita < tamanho and vetor[direita] > vetor[maior]:
            maior = direita

        if maior != raiz:
            trocar(raiz, maior)
            heapify(tamanho, maior)

    n = len(vetor)

    for i in range(n // 2 - 1, -1, -1):
        heapify(n, i)

    for i in range(n - 1, 0, -1):
        trocar(0, i)
        heapify(i, 0)

    return trocas


# ==========================
# EXECUÇÃO COM LIMITE DE TEMPO
# ==========================

def executar_algoritmo(nome_algoritmo, vetor, fila):
    try:
        inicio = time.perf_counter()

        if nome_algoritmo == "Insertion Sort":
            operacoes = insertion_sort(vetor)
        elif nome_algoritmo == "Merge Sort":
            operacoes = merge_sort(vetor)
        elif nome_algoritmo == "Heap Sort":
            operacoes = heap_sort(vetor)
        else:
            raise ValueError("Algoritmo desconhecido.")

        fim = time.perf_counter()
        tempo = fim - inicio

        if vetor != sorted(vetor):
            raise ValueError("O vetor não foi ordenado corretamente.")

        fila.put(("OK", tempo, operacoes))

    except Exception as erro:
        fila.put(("ERRO", str(erro), None))


def executar_com_timeout(nome_algoritmo, vetor_original):
    fila = mp.Queue()

    processo = mp.Process(
        target=executar_algoritmo,
        args=(nome_algoritmo, vetor_original.copy(), fila)
    )

    processo.start()
    processo.join(LIMITE_TEMPO)

    if processo.is_alive():
        processo.terminate()
        processo.join()
        return None, None, "Não finalizou em até 5 minutos"

    if not fila.empty():
        status, tempo, operacoes = fila.get()

        if status == "OK":
            return tempo, operacoes, "OK"
        else:
            return None, None, f"Erro: {tempo}"

    return None, None, "Erro desconhecido"


# ==========================
# EXPERIMENTO PRINCIPAL
# ==========================

def main():
    resultados = []

    for tamanho in TAMANHOS:
        print(f"\nGerando vetor com {tamanho} elementos...")

        gerador = random.Random(SEED + tamanho)
        vetor_original = [gerador.randint(0, tamanho * 10) for _ in range(tamanho)]

        for algoritmo in ALGORITMOS:
            print(f"\nAlgoritmo: {algoritmo} | Tamanho: {tamanho}")

            tempos = []
            operacoes_lista = []
            status_final = "OK"

            execucoes = [None, None, None]

            for repeticao in range(REPETICOES):
                print(f"Execução {repeticao + 1}...")

                tempo, operacoes, status = executar_com_timeout(algoritmo, vetor_original)

                if status != "OK":
                    print(status)
                    execucoes[repeticao] = status
                    status_final = status
                    break

                print(f"Tempo: {tempo:.6f} segundos | Operações: {operacoes}")

                execucoes[repeticao] = tempo
                tempos.append(tempo)
                operacoes_lista.append(operacoes)

            if len(tempos) == REPETICOES:
                media = statistics.mean(tempos)
                desvio = statistics.stdev(tempos)
                operacoes_final = operacoes_lista[0]
            else:
                media = None
                desvio = None
                operacoes_final = None

            resultados.append({
                "Algoritmo": algoritmo,
                "Tamanho do vetor": tamanho,
                "Execução 1 (s)": execucoes[0],
                "Execução 2 (s)": execucoes[1],
                "Execução 3 (s)": execucoes[2],
                "Tempo médio (s)": media,
                "Desvio padrão": desvio,
                "Trocas/Movimentações": operacoes_final,
                "Status": status_final
            })

    df = pd.DataFrame(resultados)

    print("\nResultados finais:")
    print(df)

    # ==========================
    # GERAR GRÁFICO
    # ==========================

    df_grafico = df.dropna(subset=["Tempo médio (s)"])

    tabela_grafico = df_grafico.pivot(
        index="Tamanho do vetor",
        columns="Algoritmo",
        values="Tempo médio (s)"
    )
    tabela_grafico = tabela_grafico.reindex(columns=ALGORITMOS)

    ax = tabela_grafico.plot(kind="line", figsize=(10, 4), linewidth=3)

    ax.set_title("Tempo médio por tamanho do vetor", fontsize=15, pad=10)
    ax.set_xlabel("Tamanho do vetor")
    ax.set_ylabel("Tempo médio em segundos")
    ax.grid(True, linestyle="--", alpha=0.4)
    ax.legend(loc="lower center", bbox_to_anchor=(0.5, -0.32), ncol=3)

    plt.tight_layout()
    plt.savefig(ARQUIVO_GRAFICO)
    plt.close()

    # ==========================
    # GERAR EXCEL
    # ==========================

    if os.path.exists(ARQUIVO_MODELO):
        shutil.copyfile(ARQUIVO_MODELO, ARQUIVO_SAIDA)
        modo = "a"
    else:
        modo = "w"

    with pd.ExcelWriter(
        ARQUIVO_SAIDA,
        engine="openpyxl",
        mode=modo,
        if_sheet_exists="replace" if modo == "a" else None
    ) as writer:
        df.to_excel(writer, sheet_name="Resultados", index=False)

    # Colocar gráfico dentro do Excel
    from openpyxl import load_workbook
    from openpyxl.drawing.image import Image as ExcelImage

    wb = load_workbook(ARQUIVO_SAIDA)

    if "Gráfico" in wb.sheetnames:
        del wb["Gráfico"]

    ws = wb.create_sheet("Gráfico")

    img = ExcelImage(ARQUIVO_GRAFICO)
    ws.add_image(img, "A1")

    wb.save(ARQUIVO_SAIDA)

    print("\nArquivo gerado com sucesso:")
    print(ARQUIVO_SAIDA)
    print("\nGráfico gerado:")
    print(ARQUIVO_GRAFICO)


if __name__ == "__main__":
    mp.freeze_support()
    main()
